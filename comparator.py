import PySimpleGUI as sg
from openpyxl import load_workbook
from os import listdir
from os import path
from openpyxl.cell.read_only import EmptyCell

def dp(values):
    print('Inside dp')

    dp_file_path = values['-dpDpFile-']
    country_name = values['-dpCountryName-']
    spirit = values['-dpSpirit-']
    sheet = 'Time Periods'

    # load excel file
    wb = load_workbook(dp_file_path)

    # get the sheet
    ws = wb[sheet]

    # find the cell containing the country name
    header_row = ws.iter_cols(min_row=2, min_col=7, max_row=2)

    for a in header_row:
        if country_name in str(a[0].value):
           target_col = a[0]
           break

    if target_col:
        rows = ws.iter_rows(min_col=target_col.column, max_col=target_col.column, min_row=7)

        for r in rows:
            if spirit == r[0].value:
                sg.popup(f'The {spirit} has been found on {target_col.column_letter}{r[0].row}.')
                return
        
        sg.popup(f'Could not find {spirit} for country {country_name} in this file.')


def labeled(values):
    print('Inside labeled')

    dp_file_path = values['-labeledDpFile-']
    country_code = values['-labeledCountryCode-']
    max_row = int(values['-labeledMaxRow-'])
    folder_path = values['-labeledSearchFolder-']
    sheet = 'Time Periods'

    # load excel file
    wb = load_workbook(dp_file_path)

    # get the sheet
    ws = wb[sheet]

    # find the cell containing the country name
    header_row = ws.iter_cols(min_row=6, max_row=max_row, min_col=2, max_col=3)

    # get filenames to search in the folder
    filenames = []
    for i, e in enumerate(header_row):
        filenames = filenames + list(e)
    filtered = [i.value.replace('MarketCountryCode', country_code) for i in filenames if i.value != None]

    folder_files = listdir(folder_path)
    folder_files_without_extension = [x.split('.')[0] for x in folder_files]

    diff = set(filtered) - set(folder_files_without_extension) 

    result = f'The Following {len(diff)} files is not present in the specified folder:\n\n'

    for d in diff:
        result += f'- {d}\n'

    sg.popup(result)

def isZero(value):
    return value == '0' or value == 0

def isEmptyString(value):
    val = str(value).strip()
    return not bool(val)


def check_sheet_for_empty_zero(book, sheetname):
    sheet = book[sheetname]
    rows = list(sheet.iter_rows(min_row=7, min_col=1))

    # print(rows)

    sheetTitle = f'The following cells are INVALID on sheet {sheet.title} \n \n'
    r = ''

    for row in rows:
        for cell in row:
            if not isinstance(cell, EmptyCell):
                if isEmptyString(cell.value):
                    r += f'- Cell {cell.coordinate} is empty \n'

                elif cell.row == 7 and isZero(cell.value):
                        r += f'The not zero {cell.coordinate} is 0 \n'

    if r != '':
        sg.popup(sheetTitle + r, title=sheet.title)


def check_brand(values):
    print('Inside check_brand')

    brand_list_path = values['-brandListFile-']
    country_code = values['-countryCode-']
    banner_file_name = values['-countryCode-']
    brand_type = ""

    if values['-BrandCo-']:
        banner_file_name += '_BRANDCO_' + values['-brandYear-'] + '_Quarterly_' + values['-brandQuarter-']
        brand_type = 'BRANDCO'
    elif values['-KPI-']:
        banner_file_name += '_KPI_' + values['-brandYear-'] + '_Quarterly_' + values['-brandQuarter-']
        brand_type = 'KPI'
    elif values['-KPIYearly-']:
        banner_file_name += '_KPI_' + values['-brandYear-'] + '_Yearly'
        brand_type = 'KPIYEARLY'

    print('brand type is : ')
    print(brand_type)

    # add extension to filename
    banner_file_name += '.xlsx'
    baner_path = values['-brandSearchFolder-'] + '/' + banner_file_name

    # check if file exists
    if path.exists(baner_path): 
        # load files
        brand_list_file = load_workbook(brand_list_path, read_only=True)
        brand_baner_file = load_workbook(baner_path, read_only=True)

        brand_baner_file.sheetnames

        # sheets
        listSheet = brand_list_file[country_code]
        banerSheet = brand_baner_file['Home']

        # load brand list 
        cols = listSheet.iter_rows(min_row=2, min_col=7, max_col=7)
        brand_list = [item.value for sublist in list(cols) for item in sublist if item.value != None]

        if brand_type == "KPI":
            brand_list.append('TYPEOF. P3M category usage')
            brand_list.append('FLAGPROB + OTHERTYPE. Categories Used Last Occasion')
            brand_list.append('FLAGPROB + OTHERTYPE + CONSIDERATION. Last Occasion Category consideration')
            brand_list.append('REJECT. Rejected categories')

        if brand_type == 'KPIYEARLY':
            brand_list.append('MOCS')
            brand_list.append('BIA. IMAGE ASSOCIATION')

        # load baner 
        cols2 = banerSheet.iter_rows(min_row=2, min_col=2, max_col=2)
        baner_list = [item.value.removeprefix('5 Cs - ') for sublist in list(cols2) for item in sublist if item.value != None]
        
        # The 5cs diff
        diff = set(brand_list) - set(baner_list)

        result = f'The following Brand Names from brand master list is not present on the Banner file.\n\n'
        for d in diff: 
            result += f'- {d}\n'

        sg.popup(result, title='Differences')

        sheetnames = brand_baner_file.sheetnames
        del sheetnames[0]

        for sheetname in sheetnames:
            check_sheet_for_empty_zero(brand_baner_file, sheetname)

    else:
        sg.popup(f'Could not find the file: \n {baner_path}', title='Banner File Not Found')